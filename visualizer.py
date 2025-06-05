"""Utility functions for analysing and exporting optimisation results."""

from __future__ import annotations

import io
import os
from datetime import datetime
from typing import Iterable

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import scipy.signal
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font


# ---------------------------------------------------------------------------
# Plotting and analysis helpers
# ---------------------------------------------------------------------------

def analyze_step1_results(
    soc_daa_hours: Iterable[float],
    cha_daa_quarters: Iterable[float],
    dis_daa_quarters: Iterable[float],
    price_list_hourly: Iterable[float],
    power_cap: float,
    n_hours: int = 24,
) -> None:
    """Visualise step 1 results.

    Creates a three-panel plot showing hourly prices with charging markers,
    the state of charge and the charge/discharge power.
    """
    charge_power_hourly = [
        np.sum(list(cha_daa_quarters)[i * 4 : (i + 1) * 4]) * power_cap / 4
        for i in range(n_hours)
    ]
    discharge_power_hourly = [
        np.sum(list(dis_daa_quarters)[i * 4 : (i + 1) * 4]) * power_cap / 4
        for i in range(n_hours)
    ]

    price_array_hourly = np.array(list(price_list_hourly)[:n_hours])
    soc_array_hourly = np.array(list(soc_daa_hours)[:n_hours])

    local_min_idx_h = scipy.signal.argrelextrema(price_array_hourly, np.less)[0]
    local_max_idx_h = scipy.signal.argrelextrema(price_array_hourly, np.greater)[0]

    fig, (ax1, ax2, ax3) = plt.subplots(3, 1, figsize=(14, 14), sharex=True)

    ax1.plot(range(n_hours), price_array_hourly, label="Preis 1h (Step1)", color="black")
    ax1.plot(local_min_idx_h, price_array_hourly[local_min_idx_h], "go", label="Minima")
    ax1.plot(local_max_idx_h, price_array_hourly[local_max_idx_h], "ro", label="Maxima")
    for i in range(n_hours):
        if charge_power_hourly[i] > 0:
            ax1.scatter(i, price_array_hourly[i], marker="^", s=100, color="green", zorder=5)
        if discharge_power_hourly[i] > 0:
            ax1.scatter(i, price_array_hourly[i], marker="v", s=100, color="red", zorder=5)
    ax1.set_ylabel("Preis (€/kWh)")
    ax1.set_title("Step1: Börsenpreise")
    ax1.legend()
    ax1.grid(True)

    ax2.bar(range(n_hours), soc_array_hourly, label="State of Charge", color="tab:blue")
    ax2.set_ylabel("State of Charge (kWh)")
    ax2.set_title("Step1: Batterieladestand")
    ax2.grid(True)

    ax3.bar(range(n_hours), charge_power_hourly, width=0.5, label="Ladeleistung", color="green")
    ax3.bar(range(n_hours), [-x for x in discharge_power_hourly], width=0.5, label="Entladeleistung", color="red")
    ax3.axhline(0, color="black", linewidth=0.8)
    ax3.set_xlabel("Stunden")
    ax3.set_ylabel("Leistung (kW)")
    ax3.grid(True)

    plt.tight_layout()
    plt.show()


def calculate_real_cycles(
    cha_quarters: Iterable[float],
    dis_quarters: Iterable[float],
    power_cap: float,
    n_hours: int,
    min_soc: float,
    max_soc: float,
    allowed_cycles: float,
    eta_cha: float,
    eta_dis: float,
) -> float:
    """Return the actually driven cycles."""
    energy_charged = np.sum(list(cha_quarters)) * eta_cha * power_cap / 4
    energy_discharged = np.sum(list(dis_quarters)) / eta_dis * power_cap / 4
    total_energy_moved = energy_charged + energy_discharged
    usable_capacity = max_soc - min_soc
    real_cycles = total_energy_moved / (2 * usable_capacity)
    print(f"Cycles: allowed: {allowed_cycles * n_hours / 24:.2f}, real: {real_cycles:.2f}")
    return real_cycles


def combine_charge_discharge(
    cha_quarters: Iterable[float],
    dis_quarters: Iterable[float],
    power_cap: float,
) -> list[float]:
    """Combine charge and discharge lists into one energy profile (kWh)."""
    energy_profile = [
        (cha - dis) * (power_cap / 4) for cha, dis in zip(cha_quarters, dis_quarters)
    ]
    total_charged = sum(e for e in energy_profile if e > 0)
    total_discharged = sum(-e for e in energy_profile if e < 0)
    print(f"🔋 Gesamte geladen: {total_charged:.2f} kWh")
    print(f"🔻 Gesamte entladen: {total_discharged:.2f} kWh")
    return energy_profile


def auswertung_transaktionen_stuendlich(
    cha_daa_h: Iterable[float],
    dis_daa_h: Iterable[float],
    price_list_hourly: Iterable[float],
    power_cap: float,
) -> pd.DataFrame:
    """Return a DataFrame summarising hourly transactions."""
    daten = [
        {
            "Stunde": stunde,
            "Preis (€/kWh)": preis,
            "Geladen (kWh)": cha_pu * power_cap,
            "Entladen (kWh)": dis_pu * power_cap,
            "Kosten (€)": cha_pu * power_cap * preis,
            "Einnahmen (€)": dis_pu * power_cap * preis,
            "Profit (€)": dis_pu * power_cap * preis - cha_pu * power_cap * preis,
        }
        for stunde, (cha_pu, dis_pu, preis) in enumerate(
            zip(cha_daa_h, dis_daa_h, price_list_hourly), start=1
        )
        if cha_pu * power_cap > 0 or dis_pu * power_cap > 0
    ]
    df = pd.DataFrame(daten)
    df.loc["Summe"] = df[["Geladen (kWh)", "Entladen (kWh)", "Kosten (€)", "Einnahmen (€)", "Profit (€)"]].sum()
    print(df)
    return df


# ---------------------------------------------------------------------------
# Export functions
# ---------------------------------------------------------------------------

def export_full_results_to_excel_premium(
    soc_hours: Iterable[float],
    cha_quarters: Iterable[float],
    dis_quarters: Iterable[float],
    energy_profile: Iterable[float],
    profit: float,
    n_days: int,
    power_cap: float,
    energy_cap: float,
    min_soc: float,
    max_soc: float,
    eta_cha: float,
    eta_dis: float,
    n_cycles: int,
    cha_daa_h: Iterable[float],
    dis_daa_h: Iterable[float],
    folder_path: str,
    price_list_hourly: Iterable[float],
    c_rate: float,
    price_list_quarter: Iterable[float],
) -> None:
    """Export results and input parameters to an Excel file."""
    os.makedirs(folder_path, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"results_{timestamp}_{n_days}days.xlsx"
    full_path = os.path.join(folder_path, filename)

    df_inputs = pd.DataFrame(
        {
            "Parameter": [
                "power_cap (kW)",
                "energy_cap (kWh)",
                "min_soc (kWh)",
                "max_soc (kWh)",
                "eta_cha",
                "eta_dis",
                "n_cycles",
                "n_days",
                "C-Rate",
            ],
            "Wert": [
                power_cap,
                energy_cap,
                min_soc,
                max_soc,
                eta_cha,
                eta_dis,
                n_cycles,
                n_days,
                c_rate,
            ],
        }
    )

    min_len = min(len(list(soc_hours)), len(list(price_list_hourly)), len(list(cha_daa_h)), len(list(dis_daa_h)))
    df_soc = pd.DataFrame(
        {
            "Stunde": range(1, min_len + 1),
            "State of Charge (kWh)": list(soc_hours)[:min_len],
            "Preis Stunde": list(price_list_hourly)[:min_len],
            "Ladeleistung (p.u.)": list(cha_daa_h)[:min_len],
            "Entladeleistung (p.u.)": list(dis_daa_h)[:min_len],
        }
    )

    min_len_q = min(len(list(energy_profile)), len(list(cha_quarters)), len(list(dis_quarters)), len(list(price_list_quarter)))
    df_energy = pd.DataFrame(
        {
            "Viertelstunde": range(1, min_len_q + 1),
            "Lade/Entladeenergie (kWh)": list(energy_profile)[:min_len_q],
            "Ladeleistung (p.u.)": list(cha_quarters)[:min_len_q],
            "Entladeleistung (p.u.)": list(dis_quarters)[:min_len_q],
            "Preis": list(price_list_quarter)[:min_len_q],
        }
    )

    df_profit = pd.DataFrame({"Profit (€)": [profit]})

    with pd.ExcelWriter(full_path, engine="openpyxl") as writer:
        df_inputs.to_excel(writer, sheet_name="Eingabedaten", index=False)
        df_soc.to_excel(writer, sheet_name="State_of_Charge", index=False)
        df_energy.to_excel(writer, sheet_name="Energy_Profile", index=False)
        df_profit.to_excel(writer, sheet_name="Profit", index=False)

    img_path_soc = os.path.join(folder_path, "soc_plot.png")
    img_path_energy = os.path.join(folder_path, "energy_plot.png")

    plt.figure(figsize=(10, 4))
    plt.bar(range(len(list(soc_hours))), list(soc_hours), label="SoC", color="blue")
    plt.title("State of Charge Verlauf")
    plt.xlabel("Stunde")
    plt.ylabel("SoC (kWh)")
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(img_path_soc)
    plt.close()

    plt.figure(figsize=(10, 4))
    plt.bar(range(len(list(energy_profile))), list(energy_profile), label="Lade/Entladeenergie", color="purple")
    plt.axhline(0, color="black", linewidth=0.8)
    plt.title("Lade-/Entladeprofil")
    plt.xlabel("Viertelstunde")
    plt.ylabel("Energie (kWh)")
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(img_path_energy)
    plt.close()

    wb = load_workbook(full_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for column_cells in ws.columns:
            column = column_cells[0].column_letter
            max_length = max((len(str(c.value)) for c in column_cells if c.value), default=0)
            ws.column_dimensions[column].width = max_length + 2

    ws_chart = wb.create_sheet("Diagramme")
    img1 = Image(img_path_soc)
    img2 = Image(img_path_energy)
    img1.width = img2.width = 800
    img1.height = img2.height = 400
    ws_chart.add_image(img1, "A1")
    ws_chart.add_image(img2, "A30")
    wb.save(full_path)

    os.remove(img_path_soc)
    os.remove(img_path_energy)
    print(f"✅ Premium-Excel erfolgreich gespeichert unter: {full_path}")


def export_all_time_series_with_charts(
    soc_q: Iterable[float],
    cha_daa_quarters: Iterable[float],
    dis_daa_quarters: Iterable[float],
    cha_daa_q_real: Iterable[float],
    dis_daa_q_real: Iterable[float],
    step2_soc_ida: Iterable[float],
    cha_ida: Iterable[float],
    dis_ida: Iterable[float],
    step2_cha_ida_close: Iterable[float],
    step2_dis_ida_close: Iterable[float],
    combined_cha: Iterable[float],
    combined_dis: Iterable[float],
    price_list_daa: Iterable[float],
    price_list_ida: Iterable[float],
    folder_path: str = ".",
) -> None:
    """Export all quarter-hourly time series to Excel and create charts."""
    os.makedirs(folder_path, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"time_series_quarterly_{ts}.xlsx"
    filepath = os.path.join(folder_path, filename)

    N = len(list(cha_daa_quarters))
    price_daa_q = np.repeat(list(price_list_daa), 4)[:N]
    price_ida_q = list(price_list_ida)[:N]

    df = pd.DataFrame(
        {
            "Viertelstunde": range(1, N + 1),
            "Price_DAA_€/kWh": price_daa_q,
            "Price_IDA_€/kWh": price_ida_q,
            "SoC_DAA_kWh": list(soc_q),
            "DAA_Charge_pu": list(cha_daa_quarters),
            "DAA_Discharge_pu": list(dis_daa_quarters),
            "DAA_Charge_Real_kWh": list(cha_daa_q_real),
            "DAA_Discharge_Real_kWh": list(dis_daa_q_real),
            "SoC_IDA_kWh": list(step2_soc_ida),
            "IDA_Charge_pu": list(cha_ida),
            "IDA_Discharge_pu": list(dis_ida),
            "IDA_Close_Charge_pu": list(step2_cha_ida_close),
            "IDA_Close_Discharge_pu": list(step2_dis_ida_close),
            "Combined_Charge_pu": list(combined_cha),
            "Combined_Discharge_pu": list(combined_dis),
        }
    )

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="QuarterlyData", index=False)

    charts: list[tuple[str, io.BytesIO]] = []
    for col in df.columns:
        if col == "Viertelstunde":
            continue
        fig, ax = plt.subplots(figsize=(6, 3))
        ax.plot(df["Viertelstunde"], df[col])
        ax.set_title(col)
        ax.set_xlabel("Viertelstunde")
        ax.grid(True)
        buf = io.BytesIO()
        fig.tight_layout()
        fig.savefig(buf, format="png")
        plt.close(fig)
        buf.seek(0)
        charts.append((col, buf))

    wb = load_workbook(filepath)
    ws_chart = wb.create_sheet("Charts")
    row = 1
    for col_name, img_buf in charts:
        img = Image(img_buf)
        img.anchor = f"A{row}"
        ws_chart.add_image(img)
        row += 20
    wb.save(filepath)
    print(f"✅ Alle Zeitreihen + Diagramme exportiert nach: {filepath}")
